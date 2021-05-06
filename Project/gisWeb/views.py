from django.shortcuts import render
import pyodbc
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
from django.http import JsonResponse
from django.contrib import messages

book = Workbook()

sheet = book.active

row = ()

conn = pyodbc.connect('Driver={Sql Server};'
                      'Server=LAPTOP-8KHCS38V;'
                      'Database=SpatialDB;'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()


def hello(request):
    return render(request, 'index.html')

# 2


def insert(request):
    return render(request, 'insert.html')


def insertValue(request):

    Country = request.POST['InputCountry']
    City = request.POST['InputCity']
    Year = request.POST['InputYear']
    PM25 = request.POST['InputPm25']
    Latitude = request.POST['InputLa']
    Longtitude = request.POST['InputLong']
    Population = request.POST['Inputpop']
    Wbinc16 = request.POST['InputWbinc16']
    Region = request.POST['InputRegion']
    Conc = request.POST['InputConc']
    Color = request.POST['InputColor']
    cursor.execute("INSERT INTO AirPollutionPM25(Country, City, Year, pm25, latitude, longitude, population,wbinc16_text, Region, conc_pm25, color_pm25) VALUES ('%s','%s',%s,%s,%s,%s,%s,'%s','%s','%s','%s')" % (
        Country, City, Year, PM25, Latitude, Longtitude, Population, Wbinc16, Region, Conc, Color))

    # new dataframe with same columns
    df = pd.DataFrame({'Country': Country,
                       'City': City,
                       'Year': Year,
                       'pm25': PM25,
                       'latitude': Latitude,
                       'longitude': Longtitude,
                       'population': Population,
                       'wbinc16_text': Wbinc16,
                       'Region': Region,
                       'conc_pm25': Conc,
                       'color_pm25': Color
                       })
    writer = pd.ExcelWriter(
        'WHO_AnnualAirQuality_Database_2008_2017.xlsx', engine='openpyxl', mode='a')
    # try to open an existing workbook
    writer.book = load_workbook(
        'WHO_AnnualAirQuality_Database_2008_2017.xlsx')
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # read existing file
    reader = pd.read_excel(r'WHO_AnnualAirQuality_Database_2008_2017.xlsx')
    # write out the new sheet
    df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
    writer.close()

    messages.info(request, message='Save complete')
    return render(request, 'insert.html')


def four(request):
    return render(request, '4.html')

# 4C

def fourC(request):
    return render(request,'4C.html')

def fourCsearch(request):
    Country = request.POST['Country']
    cursor.execute(
        "SELECT DISTINCT(Country),PM25,Year FROM [SpatialDB].[dbo].[AirPollutionPM25] WHERE Country = '" + Country + "' ORDER BY Year")
    result = cursor.fetchall()

    for data in result:
        # new dataframe with same columns
        df = pd.DataFrame({'Country': [data.Country],
                           'PM25': [data.PM25],
                           'Year': [data.Year]})
        writer = pd.ExcelWriter('4c_result.xlsx', engine='openpyxl', mode='a')
    # try to open an existing workbook
        writer.book = load_workbook('4c_result.xlsx')
    # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # read existing file
        reader = pd.read_excel(r'4c_result.xlsx')
    # write out the new sheet
        df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
        writer.close()

    return render(request, '4C.html', {'queryresult': result})

# 4D

def fourD(request):
    return render(request,'4D.html')

def fourDsearch(request):
    Year = request.POST['year']
    Color = request.POST['color']
    cursor.execute(
        "SELECT SUM(Population) as SumOfPop FROM [SpatialDB].[dbo].[AirPollutionPM25] WHERE Year = "+Year+" AND color_pm25 ='"+Color+"'")

    result = cursor.fetchall()

    for sum in result:
        # new dataframe with same columns
        df = pd.DataFrame({'Year': Year,
                           'Color_PM25': Color,
                           'SumOfPopulation': sum.SumOfPop}, index=[0])
        writer = pd.ExcelWriter('4d_result.xlsx', engine='openpyxl', mode='a')
        # try to open an existing workbook
        writer.book = load_workbook('4d_result.xlsx')
        # copy existing sheets
        writer.sheets = dict((ws.title, ws)
                             for ws in writer.book.worksheets)
        # read existing file
        reader = pd.read_excel(r'4d_result.xlsx')
        # write out the new sheet
        df.to_excel(writer, index=False, header=False,
                    startrow=len(reader)+1)
        writer.close()

    return render(request, '4D.html', {'Sum': sum.SumOfPop})

# 5


def five(request):
    return render(request, '5A.html')

# 5A


def fiveA(request):
    Year = request.GET["year"]
    cursor.execute(
        "SELECT latitude, longitude FROM AirPollutionPM25 WHERE Year="+Year)
    result = cursor.fetchall()
    data = []
    for row in result:
        data.append(list(row))
    return JsonResponse({'latnlong': data})

# 5B


def fiveB(request):
    cursor.execute(
        "DECLARE @Point geography SELECT @Point = geom From [SpatialDB].[dbo].[AirPollutionPM25] WHERE city = 'Bangkok'\nSELECT DISTINCT top 50 latitude, longitude, geom.MakeValid().STDistance(@Point) AS Distance From [SpatialDB].[dbo].[AirPollutionPM25] Order By Distance ASC")
    result = cursor.fetchall()
    print(result)
    data = []
    for row in result:
        print(row)
        data.append(list(row))
    return JsonResponse({'latnlong': data})

# 5C


def fiveC(request):
    cursor.execute("")
    result = cursor.fetchall()
    data = []
    for row in result:
        data.append(list(row))
    return JsonResponse({'latnlong': data})

# 5D


def fiveD(request):
    cursor.execute("SELECT MAX(longitude) AS max_long, MAX(latitude) AS max_lati, MIN(longitude) AS min_longti, MIN(latitude) AS min_lati" +
                   "FROM AirPollutionPM25 WHERE Year=2009 AND country='Thailand' GROUP BY country, Year")
    result = cursor.fetchall()
    data = []
    for row in result:
        data.append(list(row))
    return JsonResponse({'latnlong': data})

# 5E


def fiveE(request):
    cursor.execute("SELECT distinct latitude, longitude from [SpatialDB].[dbo].[AirPollutionPM25] where country =(" +
                   "SELECT top 1 country  FROM[SpatialDB].[dbo].[AirPollutionPM25] WHERE Year=2011 GROUP BY country order by count(city) DESC) AND year=2011")
    result = cursor.fetchall()
    data = []
    for row in result:
        data.append(list(row))
    return JsonResponse({'latnlong': data})

# 5F


def fiveF(request):
    Year = request.GET["yearin"]
    cursor.execute(
        "SELECT Distinct latitude, longitude, country, city FROM AirPollutionPM25 WHERE wbinc16_text='low income' AND Year="+Year)
    result = cursor.fetchall()
    data = []
    for row in result:
        data.append(list(row))
    return JsonResponse({'latnlong': data})
